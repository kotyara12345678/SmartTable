"""
Экспорт в Excel
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any


class ExcelExporter:
    """Экспорт данных в Excel"""

    def __init__(self):
        self.default_styles = {
            'header': {
                'font': Font(bold=True, color="FFFFFF"),
                'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
                'alignment': Alignment(horizontal="center")
            },
            'default': {
                'font': Font(name="Arial", size=10),
                'alignment': Alignment(vertical="center")
            }
        }

    def export_data(self, data: List[List], filepath: str,
                    sheet_name: str = "Лист1") -> bool:
        """
        Экспорт данных в Excel
        """
        try:
            # Создаем DataFrame
            df = pd.DataFrame(data)

            # Создаем Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                # Получаем workbook и worksheet для форматирования
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Настройка ширины столбцов
                for col_idx in range(len(data[0]) if data else 0):
                    column_letter = get_column_letter(col_idx + 1)
                    max_length = 0
                    for row_idx in range(len(data)):
                        if col_idx < len(data[row_idx]):
                            cell_value = str(data[row_idx][col_idx])
                            max_length = max(max_length, len(cell_value))

                    adjusted_width = min(max(max_length + 2, 8.43), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Применяем границы ко всем ячейкам
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in worksheet.iter_rows(min_row=1, max_row=len(data),
                                               min_col=1, max_col=len(data[0]) if data else 0):
                    for cell in row:
                        cell.border = thin_border

            return True

        except Exception as e:
            print(f"Ошибка экспорта в Excel: {e}")
            return False