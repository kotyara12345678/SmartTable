"""
Экспорт в Excel
"""

import pandas as pd
from typing import List
from pysheets.src.core.workbook import Workbook

class ExcelExporter:
    """Экспортер в Excel"""

    @staticmethod
    def export_excel(workbook: Workbook, file_path: str):
        """Экспорт в Excel файл с сохранением стилей"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            # Удаляем стандартный лист
            if wb.worksheets:
                wb.remove(wb.active)
            for sheet in workbook.sheets:
                ws = wb.create_sheet(title=sheet.name)
                for row in range(sheet.rows):
                    for col in range(sheet.columns):
                        cell_obj = sheet.get_cell(row, col)
                        value = cell_obj.value if cell_obj else ""
                        ws_cell = ws.cell(row=row+1, column=col+1, value=value)
                        if cell_obj:
                            # Корректный hex-цвет (6 символов, без #)
                            def norm_hex(color, default):
                                if not color or color in ("#000000", "#FFFFFF", "000000", "FFFFFF"):
                                    return default
                                c = color.replace('#', '')
                                if len(c) == 3:
                                    c = ''.join([x*2 for x in c])
                                return c.upper().ljust(6, '0')

                            # Стиль шрифта
                            font = Font(
                                name=cell_obj.font_family or "Arial",
                                size=cell_obj.font_size or 11,
                                bold=bool(cell_obj.bold),
                                italic=bool(cell_obj.italic),
                                underline='single' if cell_obj.underline else None,
                                strike=bool(getattr(cell_obj, 'strike', False)),
                                color=norm_hex(cell_obj.text_color, "000000")
                            )
                            ws_cell.font = font
                            # Цвет фона
                            bg = norm_hex(cell_obj.background_color, None)
                            if bg and bg != "FFFFFF":
                                ws_cell.fill = PatternFill(
                                    fill_type="solid",
                                    start_color=bg,
                                    end_color=bg
                                )
                            # Выравнивание
                            align = cell_obj.alignment
                            if align == 'left':
                                ws_cell.alignment = Alignment(horizontal='left')
                            elif align == 'center':
                                ws_cell.alignment = Alignment(horizontal='center')
                            elif align == 'right':
                                ws_cell.alignment = Alignment(horizontal='right')
                # Настройка ширины столбцов
                for col in range(sheet.columns):
                    max_len = 10
                    for row in range(sheet.rows):
                        cell_obj = sheet.get_cell(row, col)
                        val = str(cell_obj.value) if cell_obj and cell_obj.value else ""
                        if len(val) > max_len:
                            max_len = len(val)
                    col_letter = get_column_letter(col+1)
                    ws.column_dimensions[col_letter].width = min(max_len, 50)
            wb.save(file_path)
        except ImportError:
            raise ImportError("Требуется установка openpyxl: pip install openpyxl")
        except Exception as e:
            raise ValueError(f"Ошибка экспорта в Excel: {str(e)}")

    @staticmethod
    def export_csv(workbook: Workbook, file_path: str, delimiter: str = ','):
        """Экспорт в CSV файл"""
        try:
            sheet = workbook.get_active_sheet()
            if not sheet:
                raise ValueError("Нет активного листа")

            data = sheet.get_data()
            df = pd.DataFrame(data)

            df.to_csv(
                file_path,
                sep=delimiter,
                index=False,
                header=False,
                encoding='utf-8'
            )

        except Exception as e:
            raise ValueError(f"Ошибка экспорта в CSV: {str(e)}")